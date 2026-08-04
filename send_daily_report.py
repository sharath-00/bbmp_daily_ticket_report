#!/usr/bin/env python3
"""
Main Executable Script: Fetches BBMP Ticket Data, Runs Analytics,
Generates HTML Email Dashboard, and Dispatches Email to Officials.
"""

import argparse
import csv
import os
import sys
import logging
from datetime import datetime, date, timedelta
import openpyxl
from config import Config
from ticket_engine import engine
from email_generator import generate_html_email_report
from mail_sender import EmailSender

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(name)s: %(message)s')
logger = logging.getLogger("SendDailyReport")

def write_tickets_to_openpyxl_sheet(ws, tickets_list, headers):
    ws.append(headers)
    for t in tickets_list:
        row = [t.get(h, "") for h in headers]
        ws.append(row)

def generate_reports_attachments(analytics: dict, filtered_tickets: list, all_context_tickets: list = None) -> list:
    """Export itemized tickets, all active open faults, and unresolved aging reports into ONE single multi-tab Excel Workbook attachment."""
    source_all_tickets = all_context_tickets if all_context_tickets is not None else filtered_tickets
    if not source_all_tickets and not filtered_tickets:
        logger.warning("No tickets to export to report attachments.")
        return []

    today_date = date.today()
    attachments = []

    # Enrich filtered tickets with aging metadata
    enriched_filtered = []
    for t in filtered_tickets:
        t_copy = dict(t)
        st_lower = (t.get("status", "Open") or "Open").lower()
        is_open_t = any(term in st_lower for term in ["open", "in progress", "pending", "assigned", "waiting"]) and not any(term in st_lower for term in ["closed", "duplicate"])
        t_date = engine.parse_opened_date(t.get("ticket_opened_on", ""))
        
        if is_open_t and t_date:
            age = (today_date - t_date).days
            t_copy["days_unresolved"] = age
            t_copy["unresolved_over_7_days"] = "YES" if age > 7 else "NO"
            t_copy["unresolved_over_30_days"] = "YES" if age > 30 else "NO"
            if age > 30:
                t_copy["aging_category"] = "> 30 Days"
            elif age > 7:
                t_copy["aging_category"] = "7 - 30 Days"
            else:
                t_copy["aging_category"] = "< 7 Days"
        else:
            t_copy["days_unresolved"] = "N/A (Closed)" if not is_open_t else "N/A"
            t_copy["aging_category"] = "Closed" if not is_open_t else "< 7 Days"
            t_copy["unresolved_over_7_days"] = "NO"
            t_copy["unresolved_over_30_days"] = "NO"
            
        enriched_filtered.append(t_copy)

    # Build list of ALL active open fault tickets across system from the starting (excluding Auto & Closed)
    all_open_faults = []
    for t in source_all_tickets:
        st_lower = (t.get("status", "Open") or "Open").lower()
        is_open_t = any(term in st_lower for term in ["open", "in progress", "pending", "assigned", "waiting"]) and not any(term in st_lower for term in ["closed", "duplicate"])
        if is_open_t:
            t_copy = dict(t)
            t_date = engine.parse_opened_date(t.get("ticket_opened_on", ""))
            if t_date:
                age = (today_date - t_date).days
                t_copy["days_unresolved"] = age
                if age > 30:
                    t_copy["aging_category"] = "> 30 Days"
                elif age > 7:
                    t_copy["aging_category"] = "7 - 30 Days"
                else:
                    t_copy["aging_category"] = "< 7 Days"
            else:
                t_copy["days_unresolved"] = 0
                t_copy["aging_category"] = "< 7 Days"
            all_open_faults.append(t_copy)

    # Sort all open faults by longest pending age descending
    all_open_faults.sort(key=lambda x: x.get("days_unresolved", 0) if isinstance(x.get("days_unresolved"), int) else -1, reverse=True)

    unresolved_over_7 = analytics.get("unresolved_over_7_days_tickets", [])
    unresolved_over_30 = analytics.get("unresolved_over_30_days_tickets", [])

    headers = [
        "ticket_id", "status", "ticket_opened_on", "days_unresolved", "aging_category",
        "unresolved_over_7_days", "unresolved_over_30_days", "ticket_closed_on",
        "zone", "ward", "region", "complainee", "problem_type",
        "priority", "assignee", "entity_name", "location"
    ]

    unresolved_headers = [
        "ticket_id", "status", "ticket_opened_on", "days_unresolved", "aging_category",
        "zone", "ward", "region", "complainee", "problem_type",
        "priority", "assignee", "entity_name", "location"
    ]

    # Single Multi-Tab Excel Workbook (.xlsx)
    excel_file = os.path.abspath("bbmp_ticket_analytics_report.xlsx")
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        
        # Sheet 1: Selected Period Tickets (or All Tickets)
        ws_all = wb.active
        ws_all.title = "All Tickets" if len(enriched_filtered) == len(source_all_tickets) else "Period Tickets"
        write_tickets_to_openpyxl_sheet(ws_all, enriched_filtered, headers)

        # Sheet 2: All Active Open Faults
        if all_open_faults:
            ws_open = wb.create_sheet(title=f"All Open Faults ({len(all_open_faults)})")
            write_tickets_to_openpyxl_sheet(ws_open, all_open_faults, unresolved_headers)

        # Sheet 3: Unresolved > 7 Days
        if unresolved_over_7:
            ws_u7 = wb.create_sheet(title=f"Unresolved > 7 Days ({len(unresolved_over_7)})")
            write_tickets_to_openpyxl_sheet(ws_u7, unresolved_over_7, unresolved_headers)

        # Sheet 4: Unresolved > 30 Days
        if unresolved_over_30:
            ws_u30 = wb.create_sheet(title=f"Unresolved > 30 Days ({len(unresolved_over_30)})")
            write_tickets_to_openpyxl_sheet(ws_u30, unresolved_over_30, unresolved_headers)

        wb.save(excel_file)
        logger.info(f"Generated Single Multi-Tab Excel Workbook attachment: {excel_file}")
        attachments.append(excel_file)
    except Exception as e:
        logger.error(f"Error generating Excel Workbook attachment: {e}")
        # Fallback to single CSV if Excel export fails
        main_csv = os.path.abspath("bbmp_daily_tickets_report.csv")
        try:
            with open(main_csv, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(enriched_filtered)
            attachments.append(main_csv)
        except Exception as csv_err:
            logger.error(f"Error writing fallback CSV: {csv_err}")

    return attachments

import json

THREAD_STATE_FILE = os.path.abspath("email_thread_state.json")

def load_thread_state() -> dict:
    if os.path.exists(THREAD_STATE_FILE):
        try:
            with open(THREAD_STATE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"Could not load thread state file: {e}")
    return {}

def save_thread_state(last_msg_id: str, thread_subject: str, references: list):
    try:
        data = {
            "last_message_id": last_msg_id,
            "thread_subject": thread_subject,
            "references": references[-10:],  # keep last 10 reference IDs
            "last_updated": datetime.now().isoformat()
        }
        with open(THREAD_STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        logger.info(f"Saved email thread state (Last Msg-ID: {last_msg_id})")
    except Exception as e:
        logger.error(f"Error saving thread state: {e}")

def main():
    parser = argparse.ArgumentParser(description="BBMP Daily Ticket Analytics & Email Dispatcher")
    parser.add_argument("-d", "--date", type=str, default=Config.DEFAULT_REPORT_PERIOD,
                        help="Report period ('today', 'yesterday', 'all', or 'YYYY-MM-DD')")
    parser.add_argument("-r", "--recipients", type=str, help="Comma-separated recipient email addresses")
    parser.add_argument("--dry-run", action="store_true", help="Generate HTML report preview without sending email")
    parser.add_argument("--no-csv", action="store_true", help="Do not attach CSV/Excel files to email")
    parser.add_argument("--use-cache", action="store_true", help="Use local disk cache instead of fetching live data from Firestore")
    parser.add_argument("--new-thread", action="store_true", help="Start a new email thread instead of continuing existing thread")

    args = parser.parse_args()

    date_label = args.date
    if date_label == "today":
        date_label = f"{date.today().strftime('%Y-%m-%d')} (Today)"
    elif date_label == "yesterday":
        date_label = f"{(date.today() - timedelta(days=1)).strftime('%Y-%m-%d')} (Yesterday)"
    elif date_label == "all":
        date_label = "All Time Accumulation"

    logger.info(f"Initiating BBMP Ticket Analytics Report for period: {date_label}")

    # 1. Fetch Tickets
    force_refresh = not args.use_cache
    tickets = engine.fetch_tickets(force_refresh=force_refresh)

    # 2. Filter Tickets for period
    filtered_tickets = engine.filter_tickets(tickets, target_date=args.date)
    logger.info(f"Processed {len(filtered_tickets)} tickets for period '{args.date}' out of {len(tickets)} total records.")

    # 3. Perform Analytics
    analytics = engine.analyze_tickets(filtered_tickets, full_context_tickets=tickets)

    # 4. Generate Attachments (Excel + CSVs)
    report_attachments = []
    if not args.no_csv:
        report_attachments = generate_reports_attachments(analytics, filtered_tickets, all_context_tickets=tickets)

    # 5. Generate HTML Email Dashboard
    html_dashboard = generate_html_email_report(analytics, date_label=date_label)

    # Save local HTML preview
    preview_file = os.path.abspath("bbmp_email_report_preview.html")
    with open(preview_file, "w", encoding="utf-8") as f:
        f.write(html_dashboard)
    logger.info(f"Saved local HTML email preview to: {preview_file}")

    # Load Email Threading State
    thread_state = load_thread_state() if not args.new_thread else {}
    last_msg_id = thread_state.get("last_message_id")
    prev_references = thread_state.get("references", [])

    if last_msg_id and not args.new_thread:
        subject = f"Re: {Config.EMAIL_SUBJECT_PREFIX} Executive Daily Analytics Digest"
        in_reply_to = last_msg_id
        references = list(prev_references)
        if last_msg_id not in references:
            references.append(last_msg_id)
        logger.info(f"Threading email onto existing conversation thread (Replying to: {in_reply_to})")
    else:
        subject = f"{Config.EMAIL_SUBJECT_PREFIX} Executive Daily Analytics Digest - {datetime.now().strftime('%d %b %Y')}"
        in_reply_to = None
        references = []
        logger.info("Starting a new email conversation thread.")

    if args.dry_run:
        logger.info("Dry-run mode active. Email was not dispatched.")
        print("\n" + "=" * 60)
        print("          DRY-RUN SUMMARY REPORT")
        print("          (Automated 'Auto' tickets excluded)")
        print("=" * 60)
        print(f" Period             : {date_label}")
        print(f" Subject Line       : {subject}")
        print(f" Thread Status      : {'Replying to thread (' + str(in_reply_to) + ')' if in_reply_to else 'New Thread'}")
        print(f" July 1 Raised      : {analytics.get('july1_total_tickets', 0)}")
        print(f" July 1 Open        : {analytics.get('july1_open_tickets', 0)} ({analytics.get('july1_open_slc_panels', 0)} SLC Panels, {analytics.get('july1_open_lamps', 0)} Lamps)")
        print(f" July 1 Closed      : {analytics.get('july1_closed_tickets', 0)}")
        print(f" July 1 Res Rate    : {analytics.get('july1_resolution_rate_percent', 0.0)}%")
        print("-" * 60)
        print(f" Unresolved > 7 Days: {analytics.get('unresolved_over_7_days_count', 0)} Tickets")
        print(f" Unresolved > 30 Days: {analytics.get('unresolved_over_30_days_count', 0)} Tickets (Critical)")
        print("-" * 60)
        print(f" Attachments        : {len(report_attachments)} files generated")
        for att in report_attachments:
            print(f"   - {os.path.basename(att)}")
        print(f" HTML Preview Path  : {preview_file}")
        print("=" * 60 + "\n")
        return

    # 6. Dispatch Email via SMTP
    recipients = [r.strip() for r in args.recipients.split(",") if r.strip()] if args.recipients else Config.RECIPIENT_EMAILS

    if not recipients:
        logger.error("No recipient emails configured in .env or provided via --recipients option!")
        sys.exit(1)

    sender = EmailSender()
    success, sent_msg_id = sender.send_email(
        recipients=recipients,
        subject=subject,
        html_content=html_dashboard,
        attachment_paths=report_attachments,
        in_reply_to=in_reply_to,
        references=references
    )

    if success:
        logger.info(f"Daily ticket report email successfully dispatched to officials! (ID: {sent_msg_id})")
        save_thread_state(sent_msg_id, subject, references + [sent_msg_id])
    else:
        logger.error("Failed to send daily ticket report email. Check SMTP configuration.")
        sys.exit(1)

if __name__ == "__main__":
    main()
